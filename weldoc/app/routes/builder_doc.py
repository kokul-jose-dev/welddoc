from flask import Blueprint, send_file
from app.database import db
from app.models.pipeline import Pipeline
from app.models.project import Project
from app.models.client import Client
from app.models.pipeline_material import PipelineMaterial
from app.models.weld import Weld
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import io
import os

builder_doc_bp = Blueprint("builder_doc", __name__)


@builder_doc_bp.route("/<int:pipeline_id>/builder-doc", methods=["GET"])
def generate_builder_doc(pipeline_id):
    pl = Pipeline.query.get_or_404(pipeline_id)
    pr = Project.query.get(pl.project_id) if pl.project_id else None
    cli = Client.query.get(pr.client_id) if pr and pr.client_id else None

    raw_materials = (
        PipelineMaterial.query.filter_by(pipeline_id=pipeline_id, archived=False)
        .order_by(PipelineMaterial.position)
        .all()
    )

    # Flatten pipeline materials for easy property access
    class MatView:
        def __init__(self, plm):
            pm = plm.project_material
            gm = pm.global_material if pm else None
            self.position = plm.position
            self.start_of_plumbing = plm.start_of_plumbing
            self.end_of_plumbing = plm.end_of_plumbing
            self.waz_no = plm.waz_no
            self.category = gm.category if gm else ""
            self.item_description = gm.item_description if gm else ""
            self.dn1 = gm.dn1 if gm else ""
            self.dn2 = gm.dn2 if gm else ""
            self.diameter = gm.diameter if gm else ""
            self.thickness = gm.thickness if gm else ""
            self.material_code = gm.material_code if gm else ""
            self.dien_no = gm.dien_no if gm else ""
            self.surface = gm.surface if gm else ""
            self.certificate = pm.certificate if pm else ""
            self.heat_no = pm.heat_no if pm else ""

    materials = [MatView(m) for m in raw_materials]
    welds = (
        Weld.query.filter_by(pipeline_id=pipeline_id, archived=False)
        .order_by(Weld.id)
        .all()
    )

    # Build weld lookup: (between_a, between_b) -> weld
    weld_map = {}
    for w in welds:
        if w.between_a and w.between_b:
            weld_map[(w.between_a, w.between_b)] = w
            weld_map[(w.between_b, w.between_a)] = w

    # Build connections per material position
    mat_connections = {}  # position -> [connected positions]
    for w in welds:
        if w.between_a and w.between_b:
            mat_connections.setdefault(w.between_a, []).append(w.between_b)
            mat_connections.setdefault(w.between_b, []).append(w.between_a)

    # Walk combined view order (start -> end)
    mat_by_pos = {m.position: m for m in materials}
    start_mat = next((m for m in materials if m.start_of_plumbing), materials[0] if materials else None)

    visited = set()
    combined_rows = []  # list of (type, data) tuples

    def walk(pos):
        if not pos or pos in visited:
            return
        visited.add(pos)
        mat = mat_by_pos.get(pos)
        if not mat:
            return

        # Find branch connections for badge
        conns = [p for p in mat_connections.get(pos, []) if p not in visited]
        branches = conns[1:] if len(conns) > 1 else []
        branch_info = ""
        for bp in branches:
            bw = weld_map.get((pos, bp))
            if bw:
                branch_info += f" ({bp}·W{bw.weld_no})"

        combined_rows.append(("material", mat, branch_info))

        # Walk to next (first unvisited connection)
        if conns:
            next_pos = conns[0]
            # Add weld between current and next
            w = weld_map.get((pos, next_pos))
            if w:
                combined_rows.append(("weld", w, None))
            walk(next_pos)

        # Walk branches
        for bp in branches:
            bw = weld_map.get((pos, bp))
            if bw and bp not in visited:
                combined_rows.append(("weld", bw, None))
                walk(bp)

    if start_mat:
        walk(start_mat.position)

    # Add any unvisited materials
    for m in materials:
        if m.position not in visited:
            combined_rows.append(("material", m, ""))

    # === Generate Excel (A-P = 16 columns) ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blatt 1"
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    sf = Font(bold=True, size=9); df = Font(size=9); bf = Font(bold=True, size=9)
    h7 = Font(bold=True, size=7); h6 = Font(bold=True, size=6)
    blue = PatternFill("solid", fgColor="B8CCE4"); grey = PatternFill("solid", fgColor="F2F2F2")
    wc = Alignment(wrap_text=True, vertical="center", horizontal="center")
    wr = Alignment(wrap_text=True, vertical="center")
    def bdr(r1,c1,r2,c2):
        for r in range(r1,r2+1):
            for c in range(c1,c2+1): ws.cell(r,c).border = thin
    def fl(r1,c1,r2,c2,f):
        for r in range(r1,r2+1):
            for c in range(c1,c2+1): ws.cell(r,c).fill = f
    for c,w in {"A":10,"B":12,"C":12,"D":10,"E":6,"F":14,"G":8,"H":7,"I":14,"J":8,"K":8,"L":8,"M":8,"N":22,"O":22,"P":18}.items():
        ws.column_dimensions[c].width = w
    # ROW 1-2 merged (taller rows)
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.merge_cells("A1:F2"); ws["A1"]="Hersteller / manufacturer: ISTinox AG"; ws["A1"].font=Font(bold=True,size=11); ws["A1"].alignment=wr
    ws.merge_cells("G1:N2"); ws["G1"]="Schweissnahtpr\u00fcfliste"; ws["G1"].font=Font(bold=True,size=16); ws["G1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.merge_cells("O1:P2"); ws["O1"]=""; ws["O1"].alignment=wc  # Company logo
    logo_path = os.path.join(os.path.dirname(__file__), '..', '..', 'image.png')
    if os.path.exists(logo_path):
        img = XlImage(logo_path)
        img.width = 130
        img.height = 45
        ws.add_image(img, 'O1')
    # ROW 3
    ws.merge_cells("A3:C3"); ws["A3"]="Auftrag - Nr. / Order No.:"; ws["A3"].font=sf
    ws.merge_cells("D3:F3"); ws["D3"]=pr.order_no if pr else ""; bdr(3,4,3,6)
    ws.merge_cells("G3:H3"); ws["G3"]="Kunde / Customer:"; ws["G3"].font=sf
    ws.merge_cells("I3:M3"); ws["I3"]=cli.name if cli else ""; bdr(3,9,3,13)
    ws.merge_cells("N3:O3"); ws["N3"]=f"Projekt / project: {pr.title if pr else ''}"; ws["N3"].font=sf
    ws["P3"]="Seite / Page: 1 von 1"; ws["P3"].font=df
    # ROW 4-5
    ws.merge_cells("A4:C5"); ws["A4"]="Rohrleitungs- / Zeichnungs Nr.\nPipeline- / Drawing No."; ws["A4"].font=sf; ws["A4"].alignment=wr
    ws.merge_cells("D4:F5"); ws["D4"]=pl.no; ws["D4"].font=bf; ws["D4"].alignment=wr; bdr(4,4,5,6)
    ws.merge_cells("G4:H5"); ws["G4"]="Schweissverfahren\nWelding procedure:"; ws["G4"].font=sf; ws["G4"].alignment=wr
    ws.merge_cells("I4:M5"); ws["I4"]=""; bdr(4,9,5,13)
    ws.merge_cells("N4:O5"); ws["N4"]="Schweisszusatzmaterial mit Chargen Nr.\nWelding additional material and batch No.:"; ws["N4"].font=sf; ws["N4"].alignment=wr
    ws["P4"]=""; ws["P5"]=""
    # ROW 6: Material headers
    fl(6,1,6,16,blue); bdr(6,1,6,16)
    ws["A6"]="Teil Nr.\nPart Nr."; ws["A6"].font=sf; ws["A6"].alignment=wc
    ws.merge_cells("B6:D6"); ws["B6"]="Beschreibung\nDescription"; ws["B6"].font=sf; ws["B6"].alignment=wc
    ws["E6"]="DN"; ws["E6"].font=sf; ws["E6"].alignment=wc
    ws["F6"]="Dimension"; ws["F6"].font=sf; ws["F6"].alignment=wc
    ws["G6"]="Material"; ws["G6"].font=sf; ws["G6"].alignment=wc
    ws["H6"]="Certificate\ntype"; ws["H6"].font=sf; ws["H6"].alignment=wc
    ws.merge_cells("I6:M6"); ws["I6"]="Oberfl\u00e4che\nSurface"; ws["I6"].font=sf; ws["I6"].alignment=wc
    ws.merge_cells("N6:O6"); ws["N6"]="Schmelzen/Probe Nr.\nHeat Number"; ws["N6"].font=sf; ws["N6"].alignment=wc
    ws["P6"]="WAZ Nummer\nAttest Number"; ws["P6"].font=sf; ws["P6"].alignment=wc
    # ROW 7: Pipe man / Welder / Tester
    ws.merge_cells("A7:D7"); ws["A7"]="Rohrschlosser / Pipe man"; ws["A7"].font=sf
    ws.merge_cells("E7:H7"); ws["E7"]="Schweisser / Welder"; ws["E7"].font=sf
    ws.merge_cells("I7:P7"); ws["I7"]="Pr\u00fcfer / Tester"; ws["I7"].font=sf
    # ROW 8-9: Weld headers
    bdr(8,1,9,16)
    ws.merge_cells("A8:A9"); ws["A8"]="Schweissnaht Nr.\nWeld seams no."; ws["A8"].font=h7; ws["A8"].alignment=wc
    ws.merge_cells("B8:C9"); ws["B8"]="Zeichnungs Nummer\nDrawing No."; ws["B8"].font=h7; ws["B8"].alignment=wc
    ws.merge_cells("D8:D9"); ws["D8"]="Wandst\u00e4rke [mm]\nThickness"; ws["D8"].font=h7; ws["D8"].alignment=wc
    ws.merge_cells("E8:E9"); ws["E8"]="Status"; ws["E8"].font=h7; ws["E8"].alignment=wc
    ws.merge_cells("F8:F9"); ws["F8"]="Schweisser Nr.\nWelder no."; ws["F8"].font=h7; ws["F8"].alignment=wc
    ws.merge_cells("G8:G9"); ws["G8"]="Datum\nDate"; ws["G8"].font=h7; ws["G8"].alignment=wc
    ws.merge_cells("H8:H9"); ws["H8"]="Signatur\nShort mark"; ws["H8"].font=h7; ws["H8"].alignment=wc
    ws.merge_cells("I8:I9"); ws["I8"]="Visuell"; ws["I8"].font=h7; ws["I8"].alignment=wc
    ws.merge_cells("J8:K8"); ws["J8"]="Endoskopie\nEndoscopy"; ws["J8"].font=h7; ws["J8"].alignment=wc
    ws["J9"]="Signatur"; ws["J9"].font=h6; ws["J9"].alignment=wc
    ws["K9"]="Report"; ws["K9"].font=h6; ws["K9"].alignment=wc
    ws.merge_cells("L8:M8"); ws["L8"]="Ferrit Test\nFerrite test"; ws["L8"].font=h7; ws["L8"].alignment=wc
    ws["L9"]="Signatur"; ws["L9"].font=h6; ws["L9"].alignment=wc
    ws["M9"]="Report"; ws["M9"].font=h6; ws["M9"].alignment=wc
    ws.merge_cells("N8:N9"); ws["N8"]="Gepr\u00fcft und akzeptiert\nDatum / Signatur\nHersteller\nManufacturer"; ws["N8"].font=h6; ws["N8"].alignment=wc
    ws.merge_cells("O8:O9"); ws["O8"]="Gepr\u00fcft und akzeptiert\nDatum / Signatur\nKunde (Optional)\nCustomer"; ws["O8"].font=h6; ws["O8"].alignment=wc
    ws.merge_cells("P8:P9"); ws["P8"]="Bemerkung\nRemarks\n(Bild Nr.)\n(Picture No.)"; ws["P8"].font=h7; ws["P8"].alignment=wc
    # Row 10 separator
    ws.row_dimensions[10].height = 4
    # === DATA ROWS ===
    row = 11
    for item_type, data, extra in combined_rows:
        if item_type == "material":
            m = data
            pos = m.position or ""
            if m.start_of_plumbing: pos = f"{m.position} (START)"
            if m.end_of_plumbing: pos = f"{m.position} (END)"
            if extra: pos += f" {extra}"
            dim = ""
            if m.diameter and m.thickness:
                dim = f"\u00d8{m.diameter.replace(' mm','').replace('mm','')}x{m.thickness.replace(' mm','').replace('mm','')}"
            elif m.diameter: dim = m.diameter
            ws.cell(row,1,pos).font=bf
            ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=4)
            ws.cell(row,2,m.item_description or m.category or "").font=df
            ws.cell(row,5,m.dn1 or "").font=df
            ws.cell(row,6,dim).font=df
            ws.cell(row,7,m.material_code or "").font=df
            ws.cell(row,8,m.certificate or "").font=df
            ws.merge_cells(start_row=row,start_column=9,end_row=row,end_column=13)
            ws.cell(row,9,m.dien_no or "").font=df
            ws.merge_cells(start_row=row,start_column=14,end_row=row,end_column=15)
            ws.cell(row,14,m.heat_no or "").font=df
            ws.cell(row,16,m.waz_no or "").font=df
            fl(row,1,row,16,blue); bdr(row,1,row,16); ws.row_dimensions[row].height = 22; row+=1
        elif item_type == "weld":
            w = data
            thk = ""
            ma = mat_by_pos.get(w.between_a)
            if ma and ma.thickness: thk = ma.thickness.replace(" mm","").replace("mm","")
            ws.cell(row,1,w.weld_no or "").font=df
            ws.merge_cells(start_row=row,start_column=2,end_row=row,end_column=3)
            ws.cell(row,2,pl.no).font=df
            ws.cell(row,4,thk).font=df
            ws.cell(row,5,w.type or "").font=df
            ws.cell(row,6,w.welder or "").font=df
            ws.cell(row,7,w.date or "").font=df
            for c in range(8,17): ws.cell(row,c,"").font=df
            bdr(row,1,row,16); ws.row_dimensions[row].height = 20; row+=1
    # Footer
    row+=1
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws.cell(row,1,"H... Handnaht / Manual weld seam").font=Font(size=7,bold=True)
    ws.cell(row,9,"o.k... In Ordnung").font=Font(size=7)
    row+=1
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws.cell(row,1,"O... Orbitalnaht / Orbital weld seam").font=Font(size=7,bold=True)
    ws.cell(row,9,"F... Fehler / Failure").font=Font(size=7)
    row+=1
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws.cell(row,1,"V... Vorfertigung / Prefabrication").font=Font(size=7,bold=True)
    ws.cell(row,9,"n.a... nicht Anwendbar / not available").font=Font(size=7)
    row+=1
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws.cell(row,1,"M... Montagenaht / Installation weld seam").font=Font(size=7,bold=True)

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{pl.no}_builder.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)
